"""XLSX extractor."""

from __future__ import annotations

from pathlib import Path

from automafile.extractors._meta import collect
from automafile.extractors.base import CorruptDocumentError, ExtractedDoc


# openpyxl DocumentProperties attributes — the OOXML core properties
# Windows 11 surfaces for spreadsheets.
_CORE_ATTRS = (
    "title", "subject", "creator", "keywords", "description", "lastModifiedBy",
    "category", "contentStatus", "identifier", "language", "version",
    "revision", "created", "modified", "lastPrinted",
)


def extract(path: Path) -> ExtractedDoc:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CorruptDocumentError("openpyxl is not installed") from exc
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise CorruptDocumentError(f"openpyxl failed for {path}: {exc}") from exc

    chunks: list[str] = []
    for ws in wb.worksheets:
        chunks.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            line = "\t".join("" if c is None else str(c) for c in row)
            if line.strip():
                chunks.append(line)

    raw_core: dict = {}
    try:
        cp = wb.properties
        for attr in _CORE_ATTRS:
            raw_core[attr] = getattr(cp, attr, None)
    except Exception:
        pass

    raw_custom: dict = {}
    try:
        custom = getattr(wb, "custom_doc_props", None)
        if custom is not None:
            for prop in custom.props:
                raw_custom[prop.name] = getattr(prop, "value", None)
    except Exception:
        pass

    wb.close()

    metadata = {
        **collect(raw_core, prefix="core_"),
        **collect(raw_custom, prefix="custom_"),
    }

    return ExtractedDoc(
        path=path,
        text="\n".join(chunks),
        format="xlsx",
        extracted_metadata=metadata,
    )
