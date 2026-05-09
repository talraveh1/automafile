"""XLSX extractor."""

from __future__ import annotations

from pathlib import Path

from dragndoc.config import get_settings
from dragndoc.extractors._caps import CapConfig, select_pages
from dragndoc.extractors._meta import collect
from dragndoc.extractors.base import CorruptDocumentError, ExtractedDoc, Section


# openpyxl DocumentProperties attributes — the OOXML core properties
# windows 11 surfaces for spreadsheets
_CORE_ATTRS = (
    "title", "subject", "creator", "keywords", "description", "lastModifiedBy",
    "category", "contentStatus", "identifier", "language", "version",
    "revision", "created", "modified", "lastPrinted",
)


def extract(path: Path) -> ExtractedDoc:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CorruptDocumentError("openpyxl is not installed") from exc
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise CorruptDocumentError(f"openpyxl failed for {path}: {exc}") from exc

    cfg = CapConfig.from_settings(get_settings())

    def _iter_sheets():
        for ws in wb.worksheets:
            chunks: list[str] = []
            char_count = 0
            for row in ws.iter_rows(values_only=True):
                line = "\t".join("" if c is None else str(c) for c in row)
                if not line.strip():
                    continue
                chunks.append(line)
                char_count += len(line) + 1
                if char_count >= cfg.per_page_chars:
                    # cap each sheet early so large workbooks do not dominate the prompt
                    break
            yield "\n".join(chunks)

    sheet_count = len(wb.worksheets)
    kept = select_pages(_iter_sheets(), cfg)
    sections = [
        Section(label=f"Sheet: {ws.title}", text=text, index=i)
        for i, (ws, text) in enumerate(zip(wb.worksheets, kept, strict=False))
    ]

    raw_core: dict = {}
    try:
        cp = wb.properties
        for attr in _CORE_ATTRS:
            raw_core[attr] = getattr(cp, attr, None)
    except Exception:
        pass

    raw_custom: dict = {}
    try:
        custom = getattr(wb, "custom_doc_props", None)
        if custom is not None:
            # custom OOXML properties often carry filing hints from business templates
            for prop in custom.props:
                raw_custom[prop.name] = getattr(prop, "value", None)
    except Exception:
        pass

    wb.close()

    metadata = {
        **collect(raw_core, prefix="core_"),
        **collect(raw_custom, prefix="custom_"),
    }

    return ExtractedDoc(
        path=path,
        sections=sections,
        total_sections=sheet_count,
        format="xlsx",
        extracted_metadata=metadata,
    )
